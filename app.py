from ultralytics import YOLO
import cv2
import sqlite3
from flask import Flask, request, render_template, redirect, url_for, jsonify, Response, flash, stream_with_context
from flask_cors import CORS
import os
import numpy as np
from datetime import datetime
from openpyxl import Workbook, load_workbook
import time
import traceback

app = Flask(__name__, template_folder='templates')
CORS(app)
app.secret_key = 'your-secret-key-change-this-in-production-12345'

# --- Configuration ---
UPLOAD_FOLDER = 'uploads'
MODELS_FOLDER = 'models'
ALLOWED_EXTENSIONS = {'mp4', 'avi', 'mov', 'mkv'}
MAX_CONTENT_LENGTH = 500 * 1024 * 1024
current_dir = os.path.dirname(os.path.abspath(__file__))

os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(MODELS_FOLDER, exist_ok=True)

# --- Global Settings State with SEPARATE thresholds ---
app_settings = {
    'live': {
        'model_name': 'best.pt',
        'conf_gun': 0.30,
        'conf_wound': 0.25,
        'iou': 0.7,
        'max_det': 100,
        'classes': [0, 1],
        'process_every_n_frames': 5
    },
    'upload': {
        'model_name': 'best.pt',
        'conf_gun': 0.35,
        'conf_wound': 0.30,
        'iou': 0.7,
        'max_det': 100,
        'classes': [0, 1],
        'process_every_n_frames': 3
    }
}

loaded_models = {}

def get_model(model_name):
    model_path = os.path.join(MODELS_FOLDER, model_name)
    
    if not os.path.exists(model_path):
        print(f"⚠️ Model {model_name} not found in {MODELS_FOLDER}.")
        print(f"⚠️ Using yolov8n.pt as fallback...")
        
        if 'yolov8n.pt' not in loaded_models:
            try:
                loaded_models['yolov8n.pt'] = YOLO('yolov8n.pt')
                print("✅ Fallback model yolov8n.pt loaded successfully")
            except Exception as e:
                print(f"❌ CRITICAL: Cannot load fallback model: {e}")
                return None
        return loaded_models['yolov8n.pt']
    
    if model_name not in loaded_models:
        print(f"✅ Loading {model_name} into memory...")
        try:
            loaded_models[model_name] = YOLO(model_path)
            print(f"✅ Successfully loaded {model_name}")
        except Exception as e:
            print(f"❌ Error loading {model_name}: {e}")
            traceback.print_exc()
            return None
    
    return loaded_models[model_name]

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

# --- Database Setup ---
def init_db():
    conn = sqlite3.connect('database.db')
    cursor = conn.cursor()
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            full_name TEXT,
            username TEXT,
            email TEXT UNIQUE,
            password TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS detection_history (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER,
            detection_type TEXT,
            file_path TEXT,
            timestamp TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (user_id) REFERENCES users (id)
        )
    """)
    conn.commit()
    conn.close()

init_db()

# --- FIXED Video Generator with proper streaming ---
def generate_frames(source, mode='live'):
    """
    Fixed video frame generator with proper buffering and streaming
    """
    print(f"\n{'='*60}")
    print(f"🎬 Starting {mode.upper()} stream")
    print(f"   Source: {source}")
    print(f"{'='*60}\n")
    
    cap = cv2.VideoCapture(source)
    
    # Set minimal buffer for live streams to reduce latency
    if mode == 'live':
        cap.set(cv2.CAP_PROP_BUFFERSIZE, 1)
    
    if not cap.isOpened():
        print(f"❌ Error opening video source: {source}")
        return
    
    # Get video properties
    fps = cap.get(cv2.CAP_PROP_FPS)
    width = int(cap.get(cv2.CAP_PROP_FRAME_WIDTH))
    height = int(cap.get(cv2.CAP_PROP_FRAME_HEIGHT))
    
    print(f"📹 Video Info: {width}x{height} @ {fps} FPS")
    
    settings = app_settings[mode]
    model = get_model(settings['model_name'])
    
    if model is None:
        print(f"❌ CRITICAL: Could not load model for {mode} mode")
        cap.release()
        return
    
    print(f"✅ Model loaded: {settings['model_name']}")
    print(f"   Gun threshold: {settings['conf_gun']}")
    print(f"   Wound threshold: {settings['conf_wound']}")
    print(f"   Frame skip: {settings['process_every_n_frames']}")
    print(f"\n🚀 Starting detection...\n")
    
    frame_count = 0
    last_processed_frame = None
    
    try:
        while True:
            success, frame = cap.read()
            
            if not success:
                print(f"\n⚠️ End of stream for {mode} mode")
                if mode == 'upload':
                    break
                else:
                    # For live camera, try to continue
                    time.sleep(0.1)
                    continue
            
            frame_count += 1
            
            # Log progress every 100 frames
            if frame_count % 100 == 0:
                print(f"📊 Processed {frame_count} frames...")
            
            # Read latest settings (allows live updates)
            settings = app_settings[mode]
            skip_rate = int(settings['process_every_n_frames'])
            current_model_name = settings['model_name']
            
            # Reload model if changed
            if current_model_name not in loaded_models:
                model = get_model(current_model_name)
                if model is None:
                    print("❌ Model reload failed, using previous model")
            
            # Process frame according to skip rate
            if frame_count % skip_rate == 0:
                try:
                    active_classes = settings['classes'] if settings['classes'] else []
                    
                    if not active_classes:
                        # No classes selected - show warning
                        annotated_frame = frame.copy()
                        cv2.putText(annotated_frame, "Detection Disabled - No Classes Selected", 
                                   (10, 30), cv2.FONT_HERSHEY_SIMPLEX, 0.7, (0, 165, 255), 2)
                        cv2.putText(annotated_frame, "Go to Settings to enable Gun or Wound detection", 
                                   (10, 60), cv2.FONT_HERSHEY_SIMPLEX, 0.6, (0, 165, 255), 2)
                        last_processed_frame = annotated_frame
                    else:
                        # Run detection with lowest threshold
                        min_conf = min(settings['conf_gun'], settings['conf_wound'])
                        
                        results = model.predict(
                            frame,
                            conf=min_conf,
                            iou=float(settings['iou']),
                            max_det=int(settings['max_det']),
                            classes=active_classes,
                            verbose=False
                        )
                        
                        # Filter by class-specific thresholds
                        result = results[0]
                        boxes = result.boxes
                        
                        if boxes is not None and len(boxes) > 0:
                            keep_indices = []
                            for i, box in enumerate(boxes):
                                cls = int(box.cls[0])
                                conf = float(box.conf[0])
                                
                                # Apply class-specific threshold
                                if cls == 0 and conf >= settings['conf_gun']:
                                    keep_indices.append(i)
                                elif cls == 1 and conf >= settings['conf_wound']:
                                    keep_indices.append(i)
                            
                            # Keep only boxes that pass threshold
                            if keep_indices:
                                result.boxes = boxes[keep_indices]
                            else:
                                result.boxes = None
                        
                        annotated_frame = result.plot()
                        last_processed_frame = annotated_frame
                    
                except Exception as e:
                    print(f"❌ Detection error on frame {frame_count}: {e}")
                    annotated_frame = frame.copy()
                    last_processed_frame = annotated_frame
            else:
                # Use last processed frame for skipped frames
                annotated_frame = last_processed_frame if last_processed_frame is not None else frame
            
            # Encode frame as JPEG with quality setting
            encode_param = [int(cv2.IMWRITE_JPEG_QUALITY), 85]
            ret, buffer = cv2.imencode('.jpg', annotated_frame, encode_param)
            
            if not ret:
                print(f"⚠️ Failed to encode frame {frame_count}")
                continue
            
            frame_bytes = buffer.tobytes()
            
            # Yield frame in multipart format
            yield (b'--frame\r\n'
                   b'Content-Type: image/jpeg\r\n\r\n' + frame_bytes + b'\r\n')
            
            # Control playback speed for uploaded videos
            if mode == 'upload':
                time.sleep(0.033)  # ~30 FPS
            else:
                time.sleep(0.001)  # Minimal delay for live
    
    except GeneratorExit:
        print(f"⚠️ Client disconnected from {mode} stream")
    except Exception as e:
        print(f"❌ Streaming error in {mode} mode: {e}")
        traceback.print_exc()
    finally:
        cap.release()
        cv2.destroyAllWindows()
        print(f"\n✅ Released video capture for {mode} mode")
        print(f"   Total frames processed: {frame_count}")
        print(f"{'='*60}\n")

# ============================================
# ROUTES
# ============================================

@app.route('/')
def home():
    return render_template('index.html')

@app.route('/about')
def about():
    return render_template('about.html')

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'GET':
        return render_template('login.html')
    
    data = request.form
    email = data.get('email', '')
    password = data.get('password', '')

    conn = sqlite3.connect('database.db')
    cursor = conn.cursor()
    cursor.execute("SELECT id, username FROM users WHERE email = ? AND password = ?", (email, password))
    user = cursor.fetchone()
    conn.close()

    if user:
        return redirect(url_for('dashboard'))
    else:
        return render_template('login.html', error="Invalid credentials")

@app.route('/register', methods=['GET', 'POST'])
def register():
    if request.method == 'GET':
        return render_template('register.html')
    
    try:
        data = request.form
        full_name = data.get('full_name', '')
        username = data.get('username', '')
        email = data.get('email', '')
        password = data.get('password', '')

        conn = sqlite3.connect('database.db')
        cursor = conn.cursor()
        cursor.execute("""
            INSERT INTO users (full_name, username, email, password)
            VALUES (?, ?, ?, ?)
        """, (full_name, username, email, password))
        conn.commit()
        conn.close()

        excel_file = 'user_details.xlsx'
        if not os.path.exists(excel_file):
            wb = Workbook()
            ws = wb.active
            ws.title = "Users"
            ws.append(["Full Name", "Username", "Email", "Password"])
            wb.save(excel_file)

        wb = load_workbook(excel_file)
        ws = wb["Users"]
        ws.append([full_name, username, email, password])
        wb.save(excel_file)

        return redirect(url_for('login'))
    except sqlite3.IntegrityError:
        return render_template('register.html', error="Email already exists")
    except Exception as e:
        print(f"Registration error: {e}")
        return render_template('register.html', error="Registration failed")

@app.route('/dashboard')
def dashboard():
    return render_template('dashboard.html')

# ============================================
# SETTINGS PAGE
# ============================================

@app.route('/settings', methods=['GET', 'POST'])
def settings():
    try:
        available_models = [f for f in os.listdir(MODELS_FOLDER) if f.endswith('.pt')]
    except Exception as e:
        print(f"Error reading models folder: {e}")
        available_models = []
    
    if not available_models:
        available_models = ['best.pt']
        flash("⚠️ No models found in /models folder. Please add .pt files there.", "warning")
    
    if request.method == 'POST':
        try:
            # === LIVE SETTINGS ===
            app_settings['live']['model_name'] = request.form.get('live_model', 'best.pt')
            app_settings['live']['conf_gun'] = float(request.form.get('live_conf_gun', 0.30))
            app_settings['live']['conf_wound'] = float(request.form.get('live_conf_wound', 0.25))
            app_settings['live']['iou'] = float(request.form.get('live_iou', 0.7))
            app_settings['live']['max_det'] = int(request.form.get('live_max_det', 100))
            app_settings['live']['process_every_n_frames'] = int(request.form.get('live_fps_skip', 5))
            
            selected_classes_live = []
            if request.form.get('live_class_0'):
                selected_classes_live.append(0)
            if request.form.get('live_class_1'):
                selected_classes_live.append(1)
            app_settings['live']['classes'] = selected_classes_live
            
            # === UPLOAD SETTINGS ===
            app_settings['upload']['model_name'] = request.form.get('upload_model', 'best.pt')
            app_settings['upload']['conf_gun'] = float(request.form.get('upload_conf_gun', 0.35))
            app_settings['upload']['conf_wound'] = float(request.form.get('upload_conf_wound', 0.30))
            app_settings['upload']['iou'] = float(request.form.get('upload_iou', 0.7))
            app_settings['upload']['max_det'] = int(request.form.get('upload_max_det', 100))
            app_settings['upload']['process_every_n_frames'] = int(request.form.get('upload_fps_skip', 3))
            
            selected_classes_upload = []
            if request.form.get('upload_class_0'):
                selected_classes_upload.append(0)
            if request.form.get('upload_class_1'):
                selected_classes_upload.append(1)
            app_settings['upload']['classes'] = selected_classes_upload
            
            if not selected_classes_live and not selected_classes_upload:
                flash("⚠️ Warning: All detections are disabled.", "warning")
            elif not selected_classes_live:
                flash("⚠️ Live detection disabled. Upload detection is active.", "warning")
            elif not selected_classes_upload:
                flash("⚠️ Upload detection disabled. Live detection is active.", "warning")
            else:
                flash("✅ Settings updated successfully! Changes apply immediately.", "success")
            
            print(f"✅ Settings updated:")
            print(f"   Live: {app_settings['live']}")
            print(f"   Upload: {app_settings['upload']}")
            
        except Exception as e:
            print(f"Error updating settings: {e}")
            traceback.print_exc()
            flash(f"❌ Error updating settings: {str(e)}", "error")
        
        return redirect(url_for('settings'))

    return render_template('settings.html', 
                           models=available_models, 
                           settings=app_settings)

# ============================================
# VIDEO DETECTION ROUTES
# ============================================

@app.route('/view_stream')
def view_stream():
    mode = request.args.get('mode')
    filename = request.args.get('filename', '')
    return render_template('result.html', mode=mode, filename=filename)

@app.route('/start_live')
def start_live():
    return redirect(url_for('view_stream', mode='live'))

@app.route('/upload_video', methods=['POST'])
def upload_video():
    if 'video' not in request.files:
        return jsonify({'error': 'No video file provided'}), 400
    
    video_file = request.files['video']
    if video_file.filename == '':
        return jsonify({'error': 'No selected file'}), 400
    
    if video_file and allowed_file(video_file.filename):
        filename = f"vid_{int(time.time())}_{video_file.filename}"
        filepath = os.path.join(UPLOAD_FOLDER, filename)
        
        try:
            video_file.save(filepath)
            print(f"✅ Video saved: {filepath}")
        except Exception as e:
            print(f"❌ Error saving video: {e}")
            return jsonify({'error': 'Failed to save video'}), 500
        
        return redirect(url_for('view_stream', mode='upload', filename=filename))
    
    return jsonify({'error': 'Invalid file type. Allowed: mp4, avi, mov, mkv'}), 400

# === FIXED VIDEO STREAMING ENDPOINTS ===

@app.route('/video_feed')
def video_feed():
    """Stream for Live Camera with proper headers"""
    return Response(
        stream_with_context(generate_frames(0, mode='live')),
        mimetype='multipart/x-mixed-replace; boundary=frame',
        headers={
            'Cache-Control': 'no-cache, no-store, must-revalidate',
            'Pragma': 'no-cache',
            'Expires': '0',
            'X-Accel-Buffering': 'no'
        }
    )

@app.route('/stream_uploaded_video/<filename>')
def stream_uploaded_video(filename):
    """Stream for Uploaded Video with proper headers"""
    filepath = os.path.join(UPLOAD_FOLDER, filename)
    
    if not os.path.exists(filepath):
        print(f"❌ Video not found: {filepath}")
        return "Video not found", 404
    
    print(f"✅ Streaming video: {filepath}")
    
    return Response(
        stream_with_context(generate_frames(filepath, mode='upload')),
        mimetype='multipart/x-mixed-replace; boundary=frame',
        headers={
            'Cache-Control': 'no-cache, no-store, must-revalidate',
            'Pragma': 'no-cache',
            'Expires': '0',
            'X-Accel-Buffering': 'no'
        }
    )

# === ERROR HANDLERS ===

@app.errorhandler(413)
def too_large(e):
    return jsonify({'error': 'File is too large (max 500MB)'}), 413

@app.errorhandler(500)
def server_error(e):
    print(f"❌ Server error: {e}")
    traceback.print_exc()
    return jsonify({'error': 'Internal server error'}), 500

# === DEBUG ROUTE (Optional) ===

@app.route('/debug/info')
def debug_info():
    """Debug endpoint to check system status"""
    info = {
        'upload_folder': os.path.abspath(UPLOAD_FOLDER),
        'models_folder': os.path.abspath(MODELS_FOLDER),
        'uploaded_files': os.listdir(UPLOAD_FOLDER) if os.path.exists(UPLOAD_FOLDER) else [],
        'models': os.listdir(MODELS_FOLDER) if os.path.exists(MODELS_FOLDER) else [],
        'loaded_models': list(loaded_models.keys()),
        'settings': app_settings
    }
    return jsonify(info)

if __name__ == '__main__':
    print("=" * 60)
    print("🚀 Starting Wildlife Protection Detection System")
    print("=" * 60)
    print(f"📁 Upload folder: {os.path.abspath(UPLOAD_FOLDER)}")
    print(f"🤖 Models folder: {os.path.abspath(MODELS_FOLDER)}")
    print(f"⚙️  Settings: http://127.0.0.1:5000/settings")
    print(f"🏠 Dashboard: http://127.0.0.1:5000/dashboard")
    print(f"🔍 Debug Info: http://127.0.0.1:5000/debug/info")
    print("=" * 60)
    app.run(debug=True, threaded=True, host='0.0.0.0', port=5000)
